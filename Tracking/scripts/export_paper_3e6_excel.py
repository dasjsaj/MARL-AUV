"""Export live/final AUV6DOF paper 3e6 results to an Excel workbook.

The long-test runner writes CSV/JSON artifacts for robustness while training is
running. This script is intentionally read-only with respect to run directories:
it scans existing artifacts and builds a compact workbook for monitoring and
paper bookkeeping.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import pickle
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_ROOTS = {
    "medium": Path("artifacts/auv6dof_stg_medium_3e6_formal_v2"),
    "hard": Path("artifacts/auv6dof_stg_hard_3e6_formal_v2"),
}
DEFAULT_OUTPUT = Path("artifacts/stg_3e6_suite_logs/paper_3e6_live_summary.xlsx")


def _read_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _read_result(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        with path.open("rb") as f:
            obj = pickle.load(f)
        return obj if isinstance(obj, dict) else {}
    except Exception as exc:  # pragma: no cover - diagnostic path
        return {"load_error": str(exc)}


def _float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        x = float(value)
    except (TypeError, ValueError):
        return None
    return x if math.isfinite(x) else None


def _last_float(row: Dict[str, Any], *names: str) -> Optional[float]:
    for name in names:
        value = _float(row.get(name))
        if value is not None:
            return value
    return None


def _pass_bool(value: Optional[float], op: str, threshold: float) -> Optional[bool]:
    if value is None:
        return None
    if op == "<":
        return value < threshold
    if op == "<=":
        return value <= threshold
    if op == ">":
        return value > threshold
    if op == ">=":
        return value >= threshold
    raise ValueError(op)


def _discover_runs(root: Path) -> Iterable[Path]:
    base = root / "auv6dof"
    if not base.exists():
        return []
    runs = [
        p
        for p in base.glob("*/seed_*/*")
        if p.is_dir() and ((p / "summary.json").exists() or (p / "config.json").exists() or (p / "eval_detail.csv").exists())
    ]
    return sorted(runs)


def _baseline_stats(root: Path) -> Dict[str, Optional[float]]:
    data = _read_json(root / "random_policy_baseline.json")
    out: Dict[str, Optional[float]] = {
        "mean_return": None,
        "mean_tracking_error": _float(data.get("mean_tracking_error")),
        "mean_target_lost": _float(data.get("mean_target_lost")),
    }
    for key in ("mean_return", "episode_return_mean", "mean_episode_return"):
        value = _float(data.get(key))
        if value is not None:
            out["mean_return"] = value
            break
    return out


def _window_mean(values: List[Optional[float]], frac: float = 0.3) -> tuple[Optional[float], Optional[float]]:
    finite = [float(v) for v in values if v is not None and math.isfinite(float(v))]
    if len(finite) < 2:
        return None, None
    n = max(1, int(math.ceil(len(finite) * frac)))
    return sum(finite[:n]) / n, sum(finite[-n:]) / n


def _auc(values: List[Optional[float]]) -> Optional[float]:
    finite = [float(v) for v in values if v is not None and math.isfinite(float(v))]
    if not finite:
        return None
    if len(finite) == 1:
        return finite[0]
    area = 0.0
    for left, right in zip(finite[:-1], finite[1:]):
        area += 0.5 * (left + right)
    return area / max(1, len(finite) - 1)


def collect_rows(roots: Dict[str, Path]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for scenario, root in roots.items():
        baseline = _baseline_stats(root)
        random_return = baseline["mean_return"]
        random_tracking_error = baseline["mean_tracking_error"]
        random_lost = baseline["mean_target_lost"]
        for run_dir in _discover_runs(root):
            rel = run_dir.relative_to(root).parts
            algo = rel[1]
            seed = rel[2].replace("seed_", "")
            eval_rows = _read_csv_rows(run_dir / "eval_detail.csv")
            latest = eval_rows[-1] if eval_rows else {}
            result = _read_result(run_dir / "exp" / "result.pkl")
            cfg_payload = _read_json(run_dir / "config.json")
            env_cfg = cfg_payload.get("contract", {}).get("env_cfg", {})
            obs_cfg = env_cfg.get("obs", {}) if isinstance(env_cfg.get("obs"), dict) else {}
            algo_profile = str(env_cfg.get("algo_profile", ""))
            semantic_enabled = bool(
                env_cfg.get("semantic_enabled", False)
                or obs_cfg.get("include_semantic_features", False)
                or obs_cfg.get("include_semantic_graph_features", False)
            )
            action_mode = str(env_cfg.get("action_control_mode", ""))
            trained = _read_json(run_dir / "trained_policy_eval.json")
            eval_returns = [_last_float(row, "eval_return") for row in eval_rows]
            eval_distances = [
                _last_float(row, "mean_target_distance", "tail100_mean_target_distance", "tail_mean_target_distance")
                for row in eval_rows
            ]
            eval_errors = [_last_float(row, "mean_tracking_error", "tail100_mean_tracking_error") for row in eval_rows]
            eval_lost = [_last_float(row, "mean_target_lost", "tail100_target_lost_rate") for row in eval_rows]
            eval_sat = [
                _last_float(row, "mean_action_saturation_rate", "tail100_action_saturation_rate")
                for row in eval_rows
            ]
            return_first, return_last = _window_mean(eval_returns)
            dist_first, dist_last = _window_mean(eval_distances)
            err_first, err_last = _window_mean(eval_errors)
            lost_first, lost_last = _window_mean(eval_lost)
            sat_first, sat_last = _window_mean(eval_sat)

            env_step = _float(result.get("env_step"))
            if env_step is None:
                # Live run fallback: use latest eval row, then known config budget
                # for progress reporting without pretending the run completed.
                env_step = _last_float(latest, "true_env_step", "train_step")
            train_iter = _float(result.get("train_iter"))
            completed = bool(result and "load_error" not in result and (run_dir / "summary.json").exists())
            unique_eval_steps = sorted({row.get("train_step", "") for row in eval_rows if row.get("train_step", "") != ""})
            eval_events = len(unique_eval_steps) if unique_eval_steps else len(eval_rows)
            tail_dist = _last_float(
                latest,
                "tail100_mean_target_distance",
                "tail_mean_target_distance",
                "mean_target_distance",
            )
            tail_err = _last_float(latest, "tail100_mean_tracking_error")
            lost = _last_float(latest, "tail100_target_lost_rate", "mean_target_lost")
            sat = _last_float(
                latest,
                "tail100_action_saturation_rate",
                "mean_action_saturation_rate",
                "mean_action_clip_rate",
            )
            tracking_reward = _last_float(latest, "mean_tracking_reward")
            control_cost = _last_float(latest, "mean_control_cost")
            eval_return = _last_float(latest, "eval_return")
            control_ratio = None
            if control_cost is not None and tracking_reward not in (None, 0):
                control_ratio = control_cost / abs(tracking_reward)

            latest_vs_random = None
            if eval_return is not None and random_return is not None:
                latest_vs_random = eval_return > random_return
            error_beats_random = None
            if err_last is not None and random_tracking_error is not None:
                error_beats_random = err_last < random_tracking_error
            lost_beats_random = None
            if lost_last is not None and random_lost is not None:
                lost_beats_random = lost_last <= random_lost
            evidence_parts = [
                len(eval_rows) >= 10,
                latest_vs_random,
                (return_last is not None and return_first is not None and return_last > return_first) or latest_vs_random,
                (err_last is not None and err_first is not None and err_last < err_first) or error_beats_random,
                dist_last is not None and dist_first is not None and dist_last < dist_first,
                (lost_last is not None and lost_first is not None and (lost_last <= lost_first or lost_last < 0.25))
                or lost_beats_random,
                sat_last is not None and sat_first is not None and (sat_last <= sat_first or sat_last <= 0.2),
                sat is not None and sat <= 0.2,
                control_ratio is not None and control_ratio <= 0.5,
            ]
            evidence_acceptance = all(x is True for x in evidence_parts)

            strict_pass_parts = [
                _pass_bool(tail_dist, "<", 0.015),
                _pass_bool(tail_err, "<", 0.005),
                _pass_bool(lost, "<=", 0.0),
                _pass_bool(sat, "<=", 0.10),
                _pass_bool(control_ratio, "<=", 0.35),
                latest_vs_random,
            ]
            strict_pass = all(x is True for x in strict_pass_parts)
            partial = any(x is None for x in strict_pass_parts)

            rows.append(
                {
                    "scenario": scenario,
                    "algo": algo,
                    "seed": int(seed),
                    "algo_profile": algo_profile,
                    "semantic_enabled": semantic_enabled,
                    "action_mode": action_mode,
                    "status": "completed" if completed else "running",
                    "env_step": env_step,
                    "train_iter": train_iter,
                    "eval_points": len(eval_rows),
                    "eval_events": eval_events,
                    "latest_true_env_step": env_step,
                    "latest_eval_index": latest.get("eval_index", ""),
                    "latest_logged_train_step": _last_float(latest, "train_step"),
                    "latest_eval_return": eval_return,
                    "eval_return_first30": return_first,
                    "eval_return_last30": return_last,
                    "eval_return_auc": _auc(eval_returns),
                    "random_mean_return": random_return,
                    "random_mean_tracking_error": random_tracking_error,
                    "random_mean_target_lost": random_lost,
                    "latest_return_gt_random": latest_vs_random,
                    "target_distance_first30": dist_first,
                    "target_distance_last30": dist_last,
                    "tracking_error_first30": err_first,
                    "tracking_error_last30": err_last,
                    "target_lost_first30": lost_first,
                    "target_lost_last30": lost_last,
                    "action_saturation_first30": sat_first,
                    "action_saturation_last30": sat_last,
                    "tail100_mean_target_distance": tail_dist,
                    "tail100_mean_tracking_error": tail_err,
                    "tail100_target_lost_rate": lost,
                    "tail100_action_saturation_rate": sat,
                    "mean_tracking_reward": tracking_reward,
                    "mean_control_cost": control_cost,
                    "control_cost_to_tracking_reward": control_ratio,
                    "distance_pass_lt_0.015": _pass_bool(tail_dist, "<", 0.015),
                    "tracking_error_pass_lt_0.005": _pass_bool(tail_err, "<", 0.005),
                    "lost_pass_eq_0": _pass_bool(lost, "<=", 0.0),
                    "saturation_pass_le_0.10": _pass_bool(sat, "<=", 0.10),
                    "control_ratio_pass_le_0.35": _pass_bool(control_ratio, "<=", 0.35),
                    "reference_threshold_pass": False if partial else strict_pass,
                    "evidence_acceptance_pass": evidence_acceptance,
                    "tracking_error_down_or_beats_random": bool(
                        (err_last is not None and err_first is not None and err_last < err_first)
                        or error_beats_random
                    ),
                    "target_distance_down": bool(
                        dist_last is not None and dist_first is not None and dist_last < dist_first
                    ),
                    "target_lost_down_or_low": bool(
                        (lost_last is not None and lost_first is not None and (lost_last <= lost_first or lost_last < 0.25))
                        or lost_beats_random
                    ),
                    "action_saturation_not_rising": bool(
                        sat_last is not None and sat_first is not None and (sat_last <= sat_first or sat_last <= 0.2)
                    ),
                    "trained_eval_mean_return": _float(trained.get("mean_return")),
                    "trained_eval_tail100_distance": _float(
                        trained.get("mean_tail100_target_distance")
                    ),
                    "trained_eval_tail100_error": _float(
                        trained.get("mean_tail100_tracking_error")
                    ),
                    "run_dir": str(run_dir),
                }
            )
    return rows


def _write_sheet(ws, rows: List[Dict[str, Any]], columns: List[str]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    pass_fill = PatternFill("solid", fgColor="E2F0D9")
    fail_fill = PatternFill("solid", fgColor="FCE4D6")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    bool_cols = {
        idx + 1
        for idx, name in enumerate(columns)
        if name.endswith("_pass")
        or name.startswith("distance_pass")
        or name.startswith("tracking_error_pass")
        or name.startswith("lost_pass")
        or name.startswith("saturation_pass")
        or name.startswith("control_ratio_pass")
        or name.endswith("_down")
        or name.endswith("_low")
        or name.endswith("_random")
        or name.endswith("_rising")
        or name == "latest_return_gt_random"
    }
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top")
            if idx in bool_cols or columns[idx - 1] == "strict_acceptance_pass":
                if cell.value is True:
                    cell.fill = pass_fill
                elif cell.value is False:
                    cell.fill = fail_fill
                else:
                    cell.fill = pending_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, name in enumerate(columns, start=1):
        width = min(max(len(name) + 2, 12), 42)
        for cell in ws[get_column_letter(idx)]:
            if cell.value is not None:
                width = min(max(width, min(len(str(cell.value)) + 2, 42)), 48)
        ws.column_dimensions[get_column_letter(idx)].width = width


def export_workbook(rows: List[Dict[str, Any]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Acceptance"

    acceptance_cols = [
        "scenario",
        "algo",
        "seed",
        "algo_profile",
        "semantic_enabled",
        "action_mode",
        "status",
        "env_step",
        "train_iter",
        "eval_points",
        "eval_events",
        "latest_true_env_step",
        "latest_eval_return",
        "eval_return_first30",
        "eval_return_last30",
        "eval_return_auc",
        "random_mean_return",
        "latest_return_gt_random",
        "target_distance_first30",
        "target_distance_last30",
        "tracking_error_first30",
        "tracking_error_last30",
        "target_lost_first30",
        "target_lost_last30",
        "tail100_mean_target_distance",
        "tail100_mean_tracking_error",
        "tail100_target_lost_rate",
        "tail100_action_saturation_rate",
        "control_cost_to_tracking_reward",
        "tracking_error_down_or_beats_random",
        "target_distance_down",
        "target_lost_down_or_low",
        "action_saturation_not_rising",
        "evidence_acceptance_pass",
        "distance_pass_lt_0.015",
        "tracking_error_pass_lt_0.005",
        "lost_pass_eq_0",
        "saturation_pass_le_0.10",
        "control_ratio_pass_le_0.35",
        "reference_threshold_pass",
    ]
    _write_sheet(ws, rows, acceptance_cols)

    ws2 = wb.create_sheet("Reward_Action")
    reward_cols = [
        "scenario",
        "algo",
        "seed",
        "algo_profile",
        "semantic_enabled",
        "action_mode",
        "status",
        "latest_eval_index",
        "latest_logged_train_step",
        "latest_eval_return",
        "mean_tracking_reward",
        "mean_control_cost",
        "control_cost_to_tracking_reward",
        "tail100_action_saturation_rate",
        "trained_eval_mean_return",
        "trained_eval_tail100_distance",
        "trained_eval_tail100_error",
    ]
    _write_sheet(ws2, rows, reward_cols)

    ws3 = wb.create_sheet("Run_Paths")
    _write_sheet(
        ws3,
        rows,
        ["scenario", "algo", "seed", "algo_profile", "semantic_enabled", "action_mode", "status", "run_dir"],
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--medium-root", type=Path, default=DEFAULT_ROOTS["medium"])
    parser.add_argument("--hard-root", type=Path, default=DEFAULT_ROOTS["hard"])
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    rows = collect_rows({"medium": args.medium_root, "hard": args.hard_root})
    export_workbook(rows, args.output)
    print(json.dumps({"output": str(args.output), "rows": len(rows)}, indent=2))


if __name__ == "__main__":
    main()
